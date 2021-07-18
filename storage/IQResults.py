# other project to be hosted here

from openpyxl import load_workbook # so we can read xlsx
import statistics # so we can find median

workbook = load_workbook("data.xlsx") # loads workbook of excel sheet
ws = workbook.active # initialises active workbook

KS3or4 = input("KS3 or 4? ")
if KS3or4 == "3":
    pass
else:
    ws = workbook["KS4"]

def main():
    print("Working with sheet {}.".format(ws))
    ks3stud = [164,
280, 311, 590, 367, 236, 549, 716, 605, 746, 330, 89, 332, 463, 449,
772, 575, 701, 750, 252]
    ks4stud = [126,245,103,67,66,177,168,5,324,176,200,319,58,15,41,6,51,201,298,33]
    # sampled students
    if KS3or4 == "3":
        print(KSeng(ks3stud),"\n")
        print(KSmath(ks3stud),"\n")
        print(KSsci(ks3stud), "\n")
        print(KSiq(ks3stud), "\n")
    else:
        print(KSeng(ks4stud),"\n")
        print(KSmath(ks4stud),"\n")
        print(KSsci(ks4stud), "\n")
        print(KSiq(ks4stud), "\n")

def KSeng(studlist):
    print("***KS3/4 ENG RESULTS***")
    engList = [] # create empty list for students
    for student in studlist: # iterate through every student
        w = ws.cell(row=student, column=25) # get student cell for subject
        engList.append(w.value) # append value to list
    print("Average is", (sum(engList) / len(engList)))
    print("Mode is", max(engList, key=engList.count))
    print("Range is", max(engList) - min(engList))
    print("Median is", statistics.median(engList))
    # output all data
    return engList # return list for further use if need be

def KSmath(studlist):
    print("***KS3/4 MATHS RESULTS***")
    mathList = []
    for student in studlist:
        w = ws.cell(row=student, column=26)
        mathList.append(w.value)
    print("Average is", (sum(mathList) / len(mathList)))
    print("Mode is", max(mathList, key=mathList.count))
    print("Range is", max(mathList) - min(mathList))
    print("Median is", statistics.median(mathList))
    return mathList

def KSsci(studlist):
    print("***KS3/4 SCIENCE RESULTS***")
    sciList = []
    for student in studlist:
        w = ws.cell(row=student, column=27)
        sciList.append(w.value)
    print("Average is", (sum(sciList) / len(sciList)))
    print("Mode is", max(sciList, key=sciList.count))
    print("Range is", max(sciList) - min(sciList))
    print("Median is", statistics.median(sciList))
    return sciList

def KSiq(studlist):
    print("*** KS3/4 IQ ***")
    iqList = []
    for student in studlist:
        w = ws.cell(row=student, column=18)
        iqList.append(w.value)
    print("Average is", (sum(iqList) / len(iqList)))
    print("Mode is", max(iqList, key=iqList.count))
    print("Range is", max(iqList) - min(iqList))
    print("Median is", statistics.median(iqList))
    return iqList

if __name__ == "__main__":
    main()
        
    

# other project to be hosted here

from openpyxl import load_workbook # so we can read xlsx
import statistics # so we can find median

def main():
    workbook = load_workbook("data.xlsx") # loads workbook of excel sheet
    ws = workbook.active # initialises active workbook
    ks3stud = [164,
280, 311, 590, 367, 236, 549, 716, 605, 746, 330, 89, 332, 463, 449,
772, 575, 701, 750, 252]
    ks4stud = [126,245,103,67,66,177,168,5,324,176,200,319,58,15,41,6,51,201,298,33]
    if workbookChoice() == "3":
        print("Working with KS3\n")
        getMaleResults(ws, ks3stud)
        getFemaleResults(ws, ks3stud)
    else:
        ws = workbook["KS4"]
        getMaleResults(ws, ks4stud)
        getFemaleResults(ws, ks4stud)
        return

def workbookChoice():
    return input("KS3 or 4? ")

def getMaleResults(ws, studentList):
    maleList = []
    for student in studentList:
        w = ws.cell(row=student, column=8)
        if w.value == "Male":
            maleList.append(student)
    print(maleList)
    maleEng = []
    maleMath = []
    maleSci = []
    for student in maleList: 
        engcell = ws.cell(row=student, column=25)
        maleEng.append(engcell.value) 
        mathcell = ws.cell(row=student, column=26)
        maleMath.append(mathcell.value)
        scicell = ws.cell(row=student, column=27)
        maleSci.append(scicell.value)
    print("There are {} males out of {} students sampled".format(len(maleList), len(studentList)))
    print("Their average maths score was {}".format(sum(maleMath) / len(maleMath)))
    print("Their average english score was {}".format(sum(maleEng) / len(maleEng)))
    print("Their average science score was {}".format(sum(maleSci) / len(maleSci)))
    print(f'{maleEng}\n{maleSci}\n{maleMath}\n')



def getFemaleResults(ws, studentList):
    femaleList = []
    for student in studentList:
        w = ws.cell(row=student, column=8)
        if w.value == "Female":
            femaleList.append(student)
    print(femaleList)
    femaleEng = []
    femaleMath = []
    femaleSci = []
    for student in femaleList: # iterate through every student
        engcell = ws.cell(row=student, column=25) # get student cell for subject
        femaleEng.append(engcell.value) # append value to list
        mathcell = ws.cell(row=student, column=26)
        femaleMath.append(mathcell.value)
        scicell = ws.cell(row=student, column=27)
        femaleSci.append(scicell.value)
    print("There are {} females out of {} students sampled".format(len(femaleList), len(studentList)))
    print("Their average maths score was {}".format(sum(femaleMath) / len(femaleMath)))
    print("Their average english score was {}".format(sum(femaleEng) / len(femaleEng)))
    print("Their average science score was {}".format(sum(femaleSci) / len(femaleSci)))
    print(f'{femaleEng}\n{femaleSci}\n{femaleMath}\n')




if __name__ == "__main__":
    main()
        
    

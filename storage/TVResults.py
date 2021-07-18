# other project to be hosted here

from openpyxl import load_workbook # so we can read xlsx
import statistics # so we can find median

def main():
    workbook = load_workbook("data.xlsx") # loads workbook of excel sheet
    ws = workbook.active # initialises active workbook
    ks3stud = [164,
280, 311, 590, 367, 236, 549, 716, 605, 746, 330, 89, 332, 463, 449,
772, 575, 701, 750, 252]
    ks4stud = [126,245,103,67,66,177,168,5,324,176,200,319,58,15,41,6,51,201,298,33] # lists of students
    if workbookChoice() == "3":
        print("Working with KS3\n")
        getTVResults(ws, ks3stud)
        getTestResults(ws, ks3stud)
    else:
        ws = workbook["KS4"]
        getTVResults(ws, ks4stud)
        getTestResults(ws, ks4stud)
        return

def workbookChoice():
    return input("KS3 or 4? ")

def getTVResults(ws, studentList):
    TVList = []
    for student in studentList:
        w = ws.cell(row=student, column=17)
        TVList.append(w.value)
    print(TVList)
    print(f"TV Average was {sum(TVList) / len(TVList)}.\n")

def getTestResults(ws, studentList):
    mathlist = []
    englist = []
    scilist = []
    for student in studentList: 
        engcell = ws.cell(row=student, column=25)
        englist.append(engcell.value) 
        mathcell = ws.cell(row=student, column=26)
        mathlist.append(mathcell.value)
        scicell = ws.cell(row=student, column=27)
        scilist.append(scicell.value)
    print(f'{mathlist}\n{englist}\n{scilist}')
    print(f"Maths avg: {sum(mathlist) / len(mathlist)}")
    print(f"Eng avg: {sum(englist) / len(englist)}")
    print(f"Sci avg: {sum(scilist) / len(scilist)}")
    


if __name__ == "__main__":
    main()
        
    
